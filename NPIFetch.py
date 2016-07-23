#!/usr/bin/python3
#
# NPIFetch.py -


import argparse
import json
import logging
from sys import exit

import openpyxl
from openpyxl.styles import PatternFill
import requests

__author__ = 'k maddux'


class NPIFetch:
    """
    Queries the npi registry and checks the data against an excel file.
    input: Excel file with the following colums:
        ProvID, Provider, NPI, Taxonomy, Sex
    output: Excel file containing any highlighted changes
    """
    def __init__(self, input, output):
        """Read the input file and write the output file.
        The input file should be in xlsx format with the following columns:
        1. Provider ID: The .1 of the providers SER
        2. The providers name in the format last, first middle
        3. The NPI number
        4. The Taxonomy code
        5. The provider sex
        Set error to True if error occurs."""
        self.output_file = output
        logging.basicConfig(filename='npi-fetch.log',level=logging.INFO)
        self.url = 'https://npiregistry.cms.hhs.gov/api/'
        self.key_prefix = ['', 'authorized_official_']
        self.redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000',
                              fill_type='solid')

        try:
            # Open the input worksheet.
            self.wb = openpyxl.load_workbook(input)
            self.sheet = self.wb.get_sheet_by_name(self.wb.sheetnames[0])
            # Make a new workbook with two worksheets.
            self.output_wb = openpyxl.Workbook()
            self.unchanged_sheet = self.output_wb.active
            self.unchanged_sheet.title = 'Unchanged records'
            self.updated_sheet = self.output_wb.create_sheet(
                                  index=1, title='Updated records')

        except Exception as e:
            # Log the exception
            print("Exception with spreadsheet: {}".format(e))
            logging.critical("Exception with spreadsheet: {}".format(e))
            exit()

        # Macros for the row indexes
        self.ID = 0
        self.NAME = 1
        self.NPI = 2
        self.TAX = 3
        self.SEX = 4

    def get_npi_data(self, params):
        try:
            r = requests.get(self.url, params=params)
            if(not r.ok):
                # TODO How do we try again a few times?
                exception_info = 'ERROR: HTTP status: {} '.format(r.status_code)
                exception_info += 'paramiters: {}'.format(params)
                print(exception_info)
                logging.critical(exception_info)
                exit()

            data = json.loads(r.text)
            # Excract only the needed data
            if(data['result_count'] == 1):
                d1 = data['results'][0]['basic']
                d1['result_count'] = data['result_count']
                d2 = data['results'][0]['taxonomies'][0]
                d2['number'] = data['results'][0]['number']
                d1.update(d2)
                return d1
            # elif(data['result_count'] == 0):
            # elif(data['result_count'] > 1):
            else:
                return data

        except KeyError as e:
            # There may be no results. Possibly an deactivated NPI.
            # Get the problem to the output sheet.
            error_info = 'KeyError in get_npi_data: {}. '.format(e)
            logging.info(error_info)
            return data

        except Exception as e:
            # Log the exception
            exception_info = 'Exception in get_npi_data(): {} '.format(e)
            exception_info += 'URL: {}'.format(self.url)
            print(exception_info)
            logging.critical(exception_info)
            exit()

    def parse_dict(self, d, provider):
        for k, v in d.items():
            if(isinstance(v, dict)):
                self.parse_dict(v, provider)
            elif(isinstance(v, list)):
                self.parse_list(v, provider)
            else:
                setattr(provider, k, v)

    def parse_list(self, l, provider):
        for v in l:
            if(isinstance(v, dict)):
                self.parse_dict(v, provider)
            elif(isinstance(v, list)):
                self.parse_list(v, provider)
            else:
                print("list item with no key: " + str(v))

    def process(self):
        """Read the intput file."""
        try:
            i = 1 # Data in the updated_sheet starts on row 1
            first_row = True
            for row in self.sheet.rows:
                if(first_row):
                    first_row = False
                else:
                    logging.info('NPI: ' + str(row[self.NPI].value))
                    # It's nice to see some output to know the program isn't stuck.
                    print('NPI: ' + str(row[self.NPI].value))
                    logging.info('SER: ' + str(row[self.ID].value))
                    logging.info('Name: ' + str(row[self.NAME].value))
                    original_row = row
                    npi = row[self.NPI].value
                    params = {'number': npi}
                    prov_data = self.get_npi_data(params)
                    mismatches = None
                    bad_results = None
                    # Make sure you got results back
                    if('result_count' not in prov_data.keys()):
                        logging.debug('no result count')
                        bad_results = 'Error: No results found! '
                        bad_results += str(prov_data)
                    elif(prov_data['result_count'] < 1):
                        logging.debug('result count < 1')
                        bad_results = 'Error: No results found! '
                        bad_results += str(prov_data)
                    elif(prov_data['result_count'] > 1):
                        logging.debug('result count > 1')
                        bad_results = 'Error: Too many results found! '
                        bad_results += str(prov_data)
                    else:
                        logging.debug('result count = 1')
                        # Validate the data. Is the name, sex, and taxonomy correct?
                        mismatches = self.xlsx_mismatches_api(prov_data, row)

                    if(mismatches):
                        print('mismatch')
                        logging.info('Mismatches: ' + mismatches)
                        # Write the row to the changed_sheet
                        self.append_row(self.updated_sheet, row)
                        self.updated_sheet.cell(row=i, column=6).value = mismatches
                        i += 1
                    elif(bad_results):
                        print('bad result')
                        self.append_row(self.updated_sheet, original_row)
                        self.updated_sheet.cell(row=i, column=6).value = bad_results
                        self.updated_sheet.cell(row=i, column=6).fill = self.redFill
                        i += 1
                    else:
                        print('unchanged')
                        # Write the row to the unchanged_sheet
                        logging.info('No change to original data')
                        self.append_row(self.unchanged_sheet, original_row)

            self.output_wb.save(self.output_file)

        except Exception as e:
            exception_info = "Exception in process(): {}".format(e)
            print(exception_info)
            logging.critical(exception_info)
            exit()

    def xlsx_mismatches_api(self, prov_data, row):
        """Checks the data for mismatches in data.
        input: a dictionary of provider data, a row object of provider data.
        output: a string with reasons for mismatch.
        """
        mismatch_info = []
        # Look for errors in the name.
        try:
            api_f_name = prov_data['first_name']
            api_l_name = prov_data['last_name']
        except KeyError as e:
            exception_info = 'KeyError Exception in '
            exception_info += 'xlsx_mismatches_api: {}\n'.format(e)
            exception_info += "Provider data: {}".format(prov_data)
            print(exception_info)
            logging.warning(exception_info)
            try:
                api_f_name = prov_data[self.key_prefix[1] + 'first_name']
                api_l_name = prov_data[self.key_prefix[1] + 'last_name']
            except KeyError as e:
                exception_info = 'KeyError Exception in '
                exception_info += 'xlsx_mismatches_api: {}\n'.format(e)
                exception_info += "Provider data: {}".format(prov_data)
                print(exception_info)
                logging.critical(exception_info)
                exit()

        try:
            xlsx_f_name, xlsx_l_name = self.parse_name(row[self.NAME].value)
            name_errors = ''

            if(len(xlsx_f_name) > 0 and len(xlsx_l_name) > 0):
                name_errors = self.name_matches(api_f_name, api_l_name,
                                                row[self.NAME].value)
                if(name_errors):
                    mismatch_info.append(name_errors)
                    row[self.NAME].value = api_l_name + ', ' + api_f_name
            else:
                mismatch_info.append('Error in SER name. Expected format: ' +
                                     'Last, First or Last, First M')
        except Exception as e:
            exception_info = "Exception with name in xlsx_mismatches_api: {}".format(e)
            print(exception_info)
            logging.critical(exception_info)
            exit()


        # Look for a blank taxonomy in the workbook.
        tax_empty = self.taxonomy_empty(row[self.TAX].value)
        if(tax_empty):
            row[self.TAX].value = prov_data['code']
            mismatch_info.append(tax_empty)
        # See if the gender is in the SER.
        try:
            xlsx_sex_empty = self.sex_empty(row[self.SEX].value, prov_data['gender'])

        except KeyError as e:
            xlsx_sex_empty = ''
            api_sex_empty = 'Error in gender: no gender info available from api'
            mismatch_info.append(api_sex_empty)
        if(xlsx_sex_empty):
            # Update the sex.
            if(prov_data['gender'] == 'F'):
                row[self.SEX].value = 'Female'
            elif(prov_data['gender'] == 'M'):
                row[self.SEX].value = 'Male'
            else:
                mismatch_info.append('Error unrecognized gender: ' +
                                      prov_data['gender'])
                mismatch_info.append(sex_empty)

        if(mismatch_info):
            return '\n'.join(mismatch_info)
        return ''

    def name_matches(self, api_f_name, api_l_name, xlsx_full_name):
        mismatch_info = ''
        if(api_f_name.upper() not in xlsx_full_name.upper()):
            mismatch_info += 'First name mismatch. NPI database returned: ' + api_f_name
            mismatch_info += '. '
            logging.warning('NPI number does not match first name')
        if(api_l_name.upper() not in xlsx_full_name.upper()):
            if(mismatch_info):
                mismatch_info += ' '
            mismatch_info += 'Last name mismatch. NPI database returned: ' + api_l_name
            mismatch_info += '. '
            logging.warning('NPI number does not match last name')
        if(mismatch_info):
            mismatch_info += 'SER contains: ' + xlsx_full_name
            mismatch_info += '. '
        return mismatch_info

    def parse_name(self, name):
        """Extracts the first and last name from the value in the name cell"""
        try:
            comas = name.count(',')
            if(comas < 1):
                raise ValueError('No coma sperating the first name ' +
                                 'from the last name: {}'.format(name))

            elif(comas > 1):
                raise ValueError('Too many comas sperating the first name ' +
                                 'from the last name: {}'.format(name))

            else:
                names = name.split(',')
                if(len(names) > 1):
                    f_name = names[1].strip()
                    f_name = self.remove_middle_name(f_name)
                    logging.info('first: '+ f_name)
                    l_name = names[0].strip()
                    logging.info('last: ' + l_name)
                    return f_name, l_name
                else:
                    raise ValueError('ERROR in parse_name. Not enough parts' +
                                     'to the name: {}'.format(name))

        except ValueError as e:
            exception_info = "Exception in parse_name: {}".format(e)
            print(exception_info)
            logging.warning(exception_info)
            return "", ""

        except Exception as e:
            exception_info = "Exception in parse_name: {}".format(e)
            print(exception_info)
            logging.critical(exception_info)
            return "", ""

    def remove_middle_name(self, f_name):
        """Returns the first name."""
        names = f_name.split(' ')
        return names[0].strip()

    def taxonomy_empty(self, xlsx_tax):
        if(xlsx_tax):
            return ''
        # if(api_tax != xlsx_tax):
            # return ('Taxonomy mismatch: {} != {}.'.format(api_tax, xlsx_tax))
        else:
            return 'Taxonomy was blank in ser.'

    def sex_empty(self, xlsx_sex, api_sex):
        if(xlsx_sex):
            return ''
        else:
            return 'No gender in SER. Gender added: ' + api_sex

    def sex_mismatch(self, api_sex, xlsx_sex):
        if(xlsx_sex and len(api_sex) > 0):
            if(xlsx_sex[0] == api_sex[0]):
                return ""
            else:
                s = 'Gender mismatch: {} != {}.'.format(api_sex, xlsx_sex)
                return s + 'Sex updated to {}'.format(api_sex)

    def append_row(self, sheet, row):
        """Append a row to the sheet."""
        try:
            cell_values = []
            for cell in row:
                cell_values.append(cell.value)
            sheet.append(cell_values)

        except Exception as e:
            exception_info = "Problem appending to xlxs file: {}".format(sheet.title)
            logging.critical(exception_info)
            print(exception_info)


if __name__ == "__main__":
    # Setup the comand line argument requirements
    parser = argparse.ArgumentParser(description='NPIFetch takes two ' +
                                'arguments: a csv input file and an ' +
                                'output file name.' +
                                NPIFetch.__doc__)

    parser.add_argument('-i', metavar='input file', type=str,
                   help='input: the name of the csv input file of ' +
                   'providers to query.', required=True)

    parser.add_argument('-o', metavar='output file', type=str,
                   help='output: the name of the file to write the ' +
                   'updated provider info to.', required=True)

    # Parse the arguments.
    cmdargs = parser.parse_args()

    npif = NPIFetch(cmdargs.i, cmdargs.o)
    npif.process()

